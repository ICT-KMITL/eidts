#!/usr/bin/python

import smtplib

from openpyxl import load_workbook
from openpyxl import Workbook

from openpyxl.workbook import Workbook

import re
EMAIL_REGEX = re.compile(r"[^@]+@[^@]+\.[^@]+")

wb = Workbook()
wb = load_workbook(filename="staff.xlsx")
# ws1 = wb.active
# print wb.sheetnames

ws1 = wb.worksheets[0]
ws2 = wb.worksheets[1]
ws3 = wb.worksheets[2]
ws4 = wb.worksheets[3]
ws5 = wb.worksheets[4]
ws6 = wb.worksheets[5]


# prerow = ws1.max_row

# print(prerow)

# thefile = open('output.txt', 'w')


# count_valid_email=0
# for i in range(prerow):
# 	email = ws1.cell(row=i+1, column=32).value
# 	if (email):
# 		if EMAIL_REGEX.match(email):
# 			thefile.write("%s\n" % email)
# 			count_valid_email+=1
# 		# print(found)

# print("count_valid_email : ",count_valid_email)

print(ws1.cell(row=3, column=2).value)
print(ws2.cell(row=3, column=2).value)
print(ws3.cell(row=3, column=2).value)
print(ws4.cell(row=3, column=2).value)
print(ws5.cell(row=3, column=2).value)
print(ws6.cell(row=3, column=2).value)
# print(ws1.cell(row=5, column=1).value)


# for i in range(49):
# 	print(ws1.cell(row=1, column=i+1).value)



# sender = 'puchong02292@gmail.com'
# receivers = ['v_hirankitti@yahoo.com','pattaya19@gmail.com','puchong02292@gmail.com']

# message = """From: From Puchong <puchong02292@gmail.com>
# To: To Visit <v_hirankitti@yahoo.com>
# Subject: Loop e-mail test

# http://iloopu.com:8000/

# This is a test e-mail message.
# """

# try:
# 	# for x in range(10):
# 	for eachuser in receivers:
# 	   # smtpObj = smtplib.SMTP('smtp.gmail.com:587')
# 	   # smtpObj.ehlo()
# 	   # smtpObj.starttls()
# 	   # smtpObj.sendmail(sender, receivers, message)         
# 	   # print "Successfully sent email"

# 	   	# SMTP_SSL Example
# 		server_ssl = smtplib.SMTP_SSL("smtp.gmail.com", 465)
# 		server_ssl.ehlo()
# 		server_ssl.login(sender, "6kqDZbt7")  
# 	   	server_ssl.sendmail(sender, eachuser, message)
# 	   	server_ssl.close()        
# 	   	print "Successfully sent email"
# except Exception, e:
#    print "Error: unable to send email"
#    raise e


